#coding=utf8
from __future__ import division
from openpyxl import Workbook, load_workbook
from openpyxl.cell import column_index_from_string

from api import *
import var as v

def processSheet1Xlsx():
    DUT = v.DUT
    SUBTYPE = ["CrashLog次数", "CrashLog用户数"]
    try:
        wb = load_workbook(v.FILE_NAME)
    except:
        print "specific file is not exist"
        return
    for cell in wb[v.SHEET1].get_cell_collection():
        if cell.value == "new":
            _ = cell.row
            newY = column_index_from_string(cell.column)
            break
    for dut in DUT:
        for cell in wb[v.SHEET1].get_cell_collection():
            if cell.value == dut:
                dutX = cell.row
                _ = column_index_from_string(cell.column)
                break
        crashCount = getKernelCrashAveDaily(dut, SUBTYPE[0])
        crashUser = getKernelCrashAveDaily(dut, SUBTYPE[1])
        if dut == DUT[-1] or dut == DUT[-2]:
            dailyActive = getClientAPPAveDailyActive(dut)
        else:
            dailyActive = getAveDailyActive(dut)
        stab = round(1-crashUser/dailyActive, 5)

        wb[v.SHEET1].cell(row=dutX, column=newY).value = crashCount
        wb[v.SHEET1].cell(row=dutX+1, column=newY).value = crashUser
        wb[v.SHEET1].cell(row=dutX+2, column=newY).value = dailyActive
        wb[v.SHEET1].cell(row=dutX+3, column=newY).value = stab
    wb.save(v.FILE_NAME)


def processSheet2Xlsx():
    DUT = v.DUT
    SUBTYPE = ["CrashLog次数", "CrashLog用户数"]
    try:
        wb = load_workbook(v.FILE_NAME)
    except:
        print "specific file is not exist"
        return
    for cell in wb[v.SHEET2].get_cell_collection():
        if cell.value == "new":
            _ = cell.row
            newY = column_index_from_string(cell.column)
            break
    for dut in DUT:
        for cell in wb[v.SHEET2].get_cell_collection():
            if cell.value == dut:
                dutX = cell.row
                _ = column_index_from_string(cell.column)
                break
        if dut == DUT[-1] or dut == DUT[-2]:
            specVersion, _ = getClientAPPVersionMostUsedDaily(dut)
            dailyActive = getClientAPPSpecVersionAveDailyActive(dut, specVersion)
        else:
            specVersion, _ = getVersionMostUsedDailyActive(dut)
            dailyActive = getSpecVersionAveDailyActive(dut, specVersion)
        crashCount = getSpecVersionKernelCrashAveDaily(dut, specVersion, SUBTYPE[0])
        crashUser = getSpecVersionKernelCrashAveDaily(dut, specVersion, SUBTYPE[1])

        stab = round(1-crashUser/dailyActive, 5)

        wb[v.SHEET2].cell(row=dutX, column=newY).value = specVersion
        wb[v.SHEET2].cell(row=dutX+1, column=newY).value = crashCount
        wb[v.SHEET2].cell(row=dutX+2, column=newY).value = crashUser
        wb[v.SHEET2].cell(row=dutX+3, column=newY).value = dailyActive
        wb[v.SHEET2].cell(row=dutX+4, column=newY).value = stab
    wb.save(v.FILE_NAME)


def processSheet3Xlsx():
    DUT = v.DUT_SUPPORT_DAEMON
    try:
        wb = load_workbook(v.FILE_NAME)
    except:
        print "specific file is not exist"
        return
    for dut in DUT:
        for cell in wb[v.SHEET3].get_cell_collection():
            if cell.value == dut:
                dutX = cell.row
                dutY = column_index_from_string(cell.column)
                break
        specVersion, dailyActive = getVersionMostUsedDailyActive(dut)
        crashTop10 = getSpecVersionDaemonCrashUserCountTop10Daily(dut, specVersion)
        for daemon in crashTop10:
            dutX = dutX + 1
            wb[v.SHEET3].cell(row=dutX, column=dutY).value = daemon[0] # 2.12.3-trafficd|11
            wb[v.SHEET3].cell(row=dutX, column=dutY+1).value = daemon[1] # count
            wb[v.SHEET3].cell(row=dutX, column=dutY+2).value = round(daemon[1]/dailyActive, 5)
    wb.save(v.FILE_NAME)


def processSheet4Xlsx():
    DUT = v.DUT
    SUBTYPE = ["CrashLog次数", "CrashLog用户数"]
    try:
        wb = load_workbook(v.FILE_NAME)
    except:
        print "specific file is not exist"
        return
    for cell in wb[v.SHEET4].get_cell_collection():
        if cell.value == "new":
            _ = cell.row
            newY = column_index_from_string(cell.column)
            break
    for dut in DUT:
        for cell in wb[v.SHEET4].get_cell_collection():
            if cell.value == dut:
                dutX = cell.row
                _ = column_index_from_string(cell.column)
                break
        if dut == DUT[-1] or dut == DUT[-2]:
            specVersion, _ = getClientAPPVersionSecondMostUsedDaily(dut)
            dailyActive = getClientAPPSpecVersionAveDailyActive(dut, specVersion)
        else:
            specVersion, _ = getVersionSecondMostUsedDailyActive(dut)
            dailyActive = getSpecVersionAveDailyActive(dut, specVersion)
        crashCount = getSpecVersionKernelCrashAveDaily(dut, specVersion, SUBTYPE[0])
        crashUser = getSpecVersionKernelCrashAveDaily(dut, specVersion, SUBTYPE[1])

        stab = round(1-crashUser/dailyActive, 5)

        wb[v.SHEET4].cell(row=dutX, column=newY).value = specVersion
        wb[v.SHEET4].cell(row=dutX+1, column=newY).value = crashCount
        wb[v.SHEET4].cell(row=dutX+2, column=newY).value = crashUser
        wb[v.SHEET4].cell(row=dutX+3, column=newY).value = dailyActive
        wb[v.SHEET4].cell(row=dutX+4, column=newY).value = stab
    wb.save(v.FILE_NAME)


# def processSheet5Xlsx():
#     DUT = v.DUT
#     SUBTYPE = ["CrashLog次数", "CrashLog用户数"]
#     try:
#         wb = load_workbook(v.FILE_NAME)
#     except:
#         print "specific file is not exist"
#         return
#     for cell in wb[v.SHEET5].get_cell_collection():
#         if cell.value == "new":
#             _ = cell.row
#             newY = column_index_from_string(cell.column)
#             break
#     for dut in DUT:
#         for cell in wb[v.SHEET5].get_cell_collection():
#             if cell.value == dut:
#                 dutX = cell.row
#                 _ = column_index_from_string(cell.column)
#                 break
#         if dut == DUT[-1] or dut == DUT[-2]:
#             specVersion, _ = getClientAPPVersionThirdMostUsedDaily(dut)
#             dailyActive = getClientAPPSpecVersionAveDailyActive(dut, specVersion)
#         else:
#             specVersion, _ = getVersionThirdMostUsedDailyActive(dut)
#             dailyActive = getSpecVersionAveDailyActive(dut, specVersion)
#         crashCount = getSpecVersionKernelCrashAveDaily(dut, specVersion, SUBTYPE[0])
#         crashUser = getSpecVersionKernelCrashAveDaily(dut, specVersion, SUBTYPE[1])
#
#         stab = round(1-crashUser/dailyActive, 5)
#
#         wb[v.SHEET5].cell(row=dutX, column=newY).value = specVersion
#         wb[v.SHEET5].cell(row=dutX+1, column=newY).value = crashCount
#         wb[v.SHEET5].cell(row=dutX+2, column=newY).value = crashUser
#         wb[v.SHEET5].cell(row=dutX+3, column=newY).value = dailyActive
#         wb[v.SHEET5].cell(row=dutX+4, column=newY).value = stab
#     wb.save(v.FILE_NAME)


def processSheet6Xlsx():
    DUT = v.DUT
    SUBTYPE = ["CrashLog次数", "CrashLog用户数"]
    try:
        wb = load_workbook(v.FILE_NAME)
    except:
        print "specific file is not exist"
        return
    for cell in wb[v.SHEET6].get_cell_collection():
        if cell.value == "new":
            _ = cell.row
            newY = column_index_from_string(cell.column)
            break
    for dut in DUT:
        for cell in wb[v.SHEET6].get_cell_collection():
            if cell.value == dut:
                dutX = cell.row
                _ = column_index_from_string(cell.column)
                break
        if dut == DUT[-1] or dut == DUT[-2]:
            specVersion, _ = getClientAPPVersionLatestStable(dut)
            dailyActive = getClientAPPSpecVersionAveDailyActive(dut, specVersion)
        else:
            specVersion, _ = getVersionLatestStableUsedDailyActive(dut)
            dailyActive = getSpecVersionAveDailyActive(dut, specVersion)
        crashCount = getSpecVersionKernelCrashAveDaily(dut, specVersion, SUBTYPE[0])
        crashUser = getSpecVersionKernelCrashAveDaily(dut, specVersion, SUBTYPE[1])

        stab = round(1-crashUser/dailyActive, 5)

        wb[v.SHEET6].cell(row=dutX, column=newY).value = specVersion
        wb[v.SHEET6].cell(row=dutX+1, column=newY).value = crashCount
        wb[v.SHEET6].cell(row=dutX+2, column=newY).value = crashUser
        wb[v.SHEET6].cell(row=dutX+3, column=newY).value = dailyActive
        wb[v.SHEET6].cell(row=dutX+4, column=newY).value = stab
    wb.save(v.FILE_NAME)


if __name__ == '__main__':
    processSheet1Xlsx()
    processSheet2Xlsx()
    processSheet3Xlsx()
    processSheet4Xlsx()
    # processSheet5Xlsx()
    processSheet6Xlsx()